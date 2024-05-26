def read_file_commands():
    with open("commands.txt", "r") as f:
        name_of_commands = f.readlines()
    name_of_commands = [command.strip().replace("\n", "").replace("\r", "").replace("\t", "") for
                        command in name_of_commands]
    return name_of_commands


def read_file_players():
    names = []
    names_of_commands = []
    numbers = []
    pols = []
    ochki_strelba = []
    ochki_beg = []

    with open("Players.txt", "r") as f:
        for line in f:
            data = line.strip().split("_")
            if len(data) >= 6:
                names.append(data[0])
                names_of_commands.append(data[1])
                numbers.append(data[2])
                pols.append(data[3])
                ochki_strelba.append(data[4])
                ochki_beg.append(data[5])

    return names, names_of_commands, numbers, pols, ochki_strelba, ochki_beg
def indexs(massiv, elem):
    x = [i for i, ltr in enumerate(massiv) if ltr == elem]

    return x

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
fontStyle = Font(size = "16")
wb = Workbook()

sheet = wb.active
sheet.merge_cells('A1:L1')
sheet.merge_cells('A2:L2')
sheet.merge_cells('A3:A4')
sheet.merge_cells('B3:B4')
sheet.merge_cells('C3:C4')
sheet.merge_cells('D3:D4')
sheet.merge_cells('E3:F3')
sheet.merge_cells('G3:H3')
sheet.merge_cells('I3:I4')
sheet.merge_cells('J3:J4')
sheet.merge_cells('K3:K4')
sheet.merge_cells('L3:L4')
sheet.cell(row=1, column=1).value = "Протокол командных результатов Спартакиады __________"
sheet.cell(row=2, column=1).value = "по служебному двоеборью в 1 группе"
sheet.cell(row=3, column=1).value = '№ п/п'
sheet.cell(row=3, column=1).alignment = Alignment(textRotation=90)
sheet.cell(row=3, column=2).value = 'Команда'
sheet.cell(row=3, column=3).value = 'Фамилия, имя'
sheet.cell(row=3, column=4).value = '№ нагрудный'
sheet.cell(row=3, column=4).alignment = Alignment(textRotation=90)
sheet.cell(row=3, column=5).value = 'Стрельба'
sheet.cell(row=4, column=5).value = 'Результат'
sheet.cell(row=4, column=5).alignment = Alignment(textRotation=90)
sheet.cell(row=4, column=6).value = 'Очки'
sheet.cell(row=4, column=6).alignment = Alignment(textRotation=90)
sheet.cell(row=3, column=7).value = 'Бег'
sheet.cell(row=4, column=7).value = 'Результат'
sheet.cell(row=4, column=7).alignment = Alignment(textRotation=90)
sheet.cell(row=4, column=8).value = 'Очки'
sheet.cell(row=4, column=8).alignment = Alignment(textRotation=90)
sheet.cell(row=3, column=9).value = 'Сумма двоеборья'
sheet.cell(row=3, column=10).value = 'Личное место'
sheet.cell(row=3, column=11).value = 'Сумма 4-х лучших личных мест'
sheet.cell(row=3, column=12).value = 'Командное место'
sheet.column_dimensions['A'].width = 5
sheet.column_dimensions['B'].width = 10
sheet.column_dimensions['C'].width = 20
sheet.column_dimensions['I'].width = 12
sheet.column_dimensions['K'].width = 15
sheet.column_dimensions['L'].width = 12
sheet.row_dimensions[1].height = 30
sheet.row_dimensions[2].height = 40
sheet.row_dimensions[4].height = 75
sheet.cell(row = 1, column = 1).font = fontStyle
sheet.cell(row = 2, column = 1).font = fontStyle
name_of_commands = read_file_commands()
names, names_of_commands, numbers, pols, result_strelba, result_beg = read_file_players()
count_of_teams = len(name_of_commands)
id_team = 1

for x in range(5, count_of_teams * 10 + 5):
    sheet.cell(row=x, column=1).value = x - 4
    if x % 10 == 5:
        y = x + 9
        sheet.merge_cells(f'B{x}:B{y}')
        sheet.cell(row=x, column=2).value = id_team
        sheet.merge_cells(f'I{x}:I{y}')
        sheet.merge_cells(f'K{x}:K{y}')
        sheet.merge_cells(f'L{x}:L{y}')
        sheet.cell(row=x, column=9).value = f"=SUM(F{x}:F{y}) + SUM(H{x}:H{y})"
        id_team += 1

for i in range(count_of_teams):
    qwe = 0
    indexes = indexs(names_of_commands, name_of_commands[i])
    for a in indexes:
        sheet.cell(row=i * 10 + 5 + qwe, column=3).value = names[a]
        sheet.cell(row=i * 10 + 5 + qwe, column=4).value = numbers[a]
        sheet.cell(row=i * 10 + 5 + qwe, column=5).value = result_strelba[a]
        sheet.cell(row=i * 10 + 5 + qwe, column=7).value = result_beg[a]
        qwe += 1


for col in sheet.columns:
    for cell in col:
        # openpyxl styles aren't mutable,
        # so you have to create a copy of the style, modify the copy, then set it back
        alignment_obj = cell.alignment.copy(horizontal='center', vertical='center', wrap_text=True)
        cell.alignment = alignment_obj

wb.save('My sheet.xlsx')
